"""Génération des exports de devis : Excel (xlsx), PDF et HTML.

Réutilise la réponse calculée par `calculate_quote` (aucun recalcul de prix ici).
Le regroupement se fait par `category` (connue du backend), équivalent des familles du front.
"""
from __future__ import annotations

import html as _html
import io
from collections import OrderedDict
from typing import Any, Optional

from .catalog import find_catalog_item
from .models import QuoteResponse

GREEN = "16A06A"
GREEN_DARK = "0F7A50"
INK = "1B2430"

LICENSE_GROUP = "Licences éditeurs"

# En-têtes de colonnes du tableau de devis (ordre = ordre d'affichage).
COLUMNS = [
    ("name", "Produit"),
    ("sku", "Référence"),
    ("quantity", "Qté"),
    ("public_unit_price", "PU public"),
    ("standard_discount_percent", "Remise cat."),
    ("discounted_unit_price", "PU remisé"),
    ("monthly_total", "Total /mois"),
    ("engagement_months", "Engagement"),
    ("engagement_total", "Total engagement"),
]


def _eur(v: float) -> str:
    s = f"{float(v or 0):,.2f}".replace(",", " ").replace(".", ",")
    return f"{s} €"


def _pct(v: float) -> str:
    v = float(v or 0)
    return "—" if v <= 0 else f"{v:.0f} %".replace(".0", "")


def _qty(v: float) -> str:
    v = float(v or 0)
    return str(int(v)) if v == int(v) else f"{v:g}"


def _category_for(sku: str, source: str) -> str:
    if source == "license":
        return LICENSE_GROUP
    item = find_catalog_item(sku)
    return (item or {}).get("category") or "Autres"


def _grouped(quote: QuoteResponse) -> "OrderedDict[str, list[Any]]":
    """Regroupe les lignes par catégorie, en préservant l'ordre d'apparition."""
    groups: "OrderedDict[str, list[Any]]" = OrderedDict()
    for line in quote.lines:
        cat = _category_for(line.sku, line.source)
        groups.setdefault(cat, []).append(line)
    return groups


def _meta(quote: QuoteResponse, meta: Optional[dict]) -> dict:
    meta = meta or {}
    return {
        "project": str(meta.get("project") or "Devis Cloud Temple"),
        "date": str(meta.get("date") or ""),
        "period_months": quote.period_months,
        "discount_percent": quote.discount_percent,
    }


def _period_label(months: int) -> str:
    if months == 1:
        return "1 mois"
    if months % 12 == 0:
        years = months // 12
        return f"{years} an" + ("s" if years > 1 else "")
    return f"{months} mois"


def _savings_breakdown(quote: QuoteResponse) -> tuple[float, float]:
    """Part mensuelle de remise catalogue (standard) et de remise commerciale."""
    std_saving = 0.0
    after_std = 0.0
    for line in quote.lines:
        pub_m = line.public_unit_price * line.quantity
        after_m = line.public_unit_price * (1 - (line.standard_discount_percent or 0) / 100) * line.quantity
        std_saving += pub_m - after_m
        after_std += after_m
    com_saving = after_std * ((quote.discount_percent or 0) / 100)
    return std_saving, com_saving


# --------------------------------------------------------------------------- #
# Excel (openpyxl)
# --------------------------------------------------------------------------- #
def quote_to_xlsx(quote: QuoteResponse, meta: Optional[dict] = None) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    m = _meta(quote, meta)
    wb = Workbook()
    ws = wb.active
    ws.title = "Devis"

    head_fill = PatternFill("solid", fgColor=GREEN)
    head_font = Font(bold=True, color="FFFFFF")
    sub_fill = PatternFill("solid", fgColor="E8F4EE")
    bold = Font(bold=True)
    right = Alignment(horizontal="right")
    thin = Side(style="thin", color="D6DEE6")
    border = Border(bottom=thin)

    ncols = len(COLUMNS)

    # En-tête document
    ws.append([f"Devis Cloud Temple — {m['project']}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color=GREEN_DARK)
    info = f"Projection : {_period_label(m['period_months'])}    Remise commerciale : {_pct(m['discount_percent'])}"
    if m["date"]:
        info = f"{m['date']}    " + info
    ws.append([info])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws.append([])

    # En-tête de tableau
    header_row = ws.max_row + 1
    ws.append([label for _, label in COLUMNS])
    for col in range(1, ncols + 1):
        c = ws.cell(row=header_row, column=col)
        c.fill = head_fill
        c.font = head_font
        c.alignment = right if col >= 3 else Alignment(horizontal="left")

    def emit_line(line) -> None:
        ws.append([
            line.name,
            line.sku,
            _qty(line.quantity),
            _eur(line.public_unit_price),
            _pct(line.standard_discount_percent),
            _eur(line.discounted_unit_price),
            _eur(line.monthly_total),
            f"{line.engagement_months} mois" if line.engagement_months > 1 else "—",
            _eur(line.engagement_total),
        ])
        r = ws.max_row
        for col in range(3, ncols + 1):
            ws.cell(row=r, column=col).alignment = right

    for cat, lines in _grouped(quote).items():
        cat_row = ws.max_row + 1
        ws.append([cat])
        ws.merge_cells(start_row=cat_row, start_column=1, end_row=cat_row, end_column=ncols)
        cc = ws.cell(row=cat_row, column=1)
        cc.fill = sub_fill
        cc.font = Font(bold=True, color=GREEN_DARK)
        subtotal = 0.0
        for line in lines:
            emit_line(line)
            subtotal += line.monthly_total
        sr = ws.max_row + 1
        ws.append(["", "", "", "", "", f"Sous-total {cat}", _eur(subtotal), "", ""])
        ws.cell(row=sr, column=6).font = bold
        ws.cell(row=sr, column=6).alignment = right
        ws.cell(row=sr, column=7).font = bold
        ws.cell(row=sr, column=7).alignment = right

    # Totaux
    ws.append([])
    std_saving, com_saving = _savings_breakdown(quote)
    totals = [
        ("Mensuel public", _eur(quote.monthly_public_total)),
        ("dont remise catalogue", "−" + _eur(std_saving)),
        ("dont remise commerciale", "−" + _eur(com_saving)),
        ("Total mensuel remisé", _eur(quote.monthly_discounted_total)),
        (f"Projection {_period_label(m['period_months'])}", _eur(quote.period_discounted_total)),
        ("Total à l'engagement", _eur(quote.total_on_engagement)),
        (f"Économie sur {_period_label(m['period_months'])}", _eur(quote.savings_total)),
    ]
    for label, value in totals:
        ws.append(["", "", "", "", "", "", "", label, value])
        r = ws.max_row
        lab = ws.cell(row=r, column=8)
        val = ws.cell(row=r, column=9)
        lab.alignment = right
        val.alignment = right
        if label.startswith("Total mensuel"):
            lab.font = Font(bold=True, size=12, color=GREEN_DARK)
            val.font = Font(bold=True, size=12, color=GREEN_DARK)

    # Largeurs de colonnes
    widths = [34, 26, 7, 12, 11, 12, 13, 12, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------------------------- #
# HTML
# --------------------------------------------------------------------------- #
def quote_to_html(quote: QuoteResponse, meta: Optional[dict] = None) -> str:
    m = _meta(quote, meta)
    e = _html.escape
    std_saving, com_saving = _savings_breakdown(quote)
    period = _period_label(m["period_months"])

    rows = []
    for cat, lines in _grouped(quote).items():
        subtotal = sum(l.monthly_total for l in lines)
        rows.append(f'<tr class="cat"><td colspan="9">{e(cat)}</td></tr>')
        for l in lines:
            eng = f"{l.engagement_months} mois" if l.engagement_months > 1 else "—"
            rows.append(
                "<tr>"
                f"<td>{e(l.name)}</td><td class='mono'>{e(l.sku)}</td>"
                f"<td class='r'>{_qty(l.quantity)}</td>"
                f"<td class='r'>{e(_eur(l.public_unit_price))}</td>"
                f"<td class='r'>{e(_pct(l.standard_discount_percent))}</td>"
                f"<td class='r'>{e(_eur(l.discounted_unit_price))}</td>"
                f"<td class='r strong'>{e(_eur(l.monthly_total))}</td>"
                f"<td class='r'>{eng}</td>"
                f"<td class='r'>{e(_eur(l.engagement_total))}</td>"
                "</tr>"
            )
        rows.append(
            f'<tr class="sub"><td colspan="6" class="r">Sous-total {e(cat)}</td>'
            f'<td class="r strong">{e(_eur(subtotal))}</td><td colspan="2"></td></tr>'
        )

    headers = "".join(
        f"<th class='{'r' if i >= 2 else ''}'>{e(label)}</th>" for i, (_, label) in enumerate(COLUMNS)
    )
    date_line = f"<span>{e(m['date'])}</span>" if m["date"] else ""
    return f"""<!doctype html>
<html lang="fr"><head><meta charset="utf-8">
<title>Devis Cloud Temple — {e(m['project'])}</title>
<style>
  * {{ box-sizing: border-box; }}
  body {{ font-family: -apple-system, "Segoe UI", Roboto, Arial, sans-serif; color: #{INK}; margin: 32px; }}
  header {{ border-bottom: 3px solid #{GREEN}; padding-bottom: 12px; margin-bottom: 20px; }}
  h1 {{ color: #{GREEN_DARK}; font-size: 22px; margin: 0 0 4px; }}
  .meta {{ color: #5b6b7d; font-size: 13px; display: flex; gap: 18px; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 12.5px; }}
  th {{ background: #{GREEN}; color: #fff; padding: 7px 9px; text-align: left; }}
  td {{ padding: 6px 9px; border-bottom: 1px solid #e3e9ef; }}
  .r {{ text-align: right; }}
  .mono {{ font-family: "SFMono-Regular", Menlo, Consolas, monospace; color: #5b6b7d; font-size: 11px; }}
  .strong {{ font-weight: 700; }}
  tr.cat td {{ background: #e8f4ee; color: #{GREEN_DARK}; font-weight: 700; }}
  tr.sub td {{ background: #f5f8fa; }}
  .totals {{ margin-top: 22px; margin-left: auto; width: 340px; }}
  .totals .row {{ display: flex; justify-content: space-between; padding: 5px 0; font-size: 13px; }}
  .totals .row.muted {{ color: #8a98a8; }}
  .totals .row.muted .v {{ text-decoration: line-through; }}
  .totals .row.detail {{ color: #5b6b7d; font-size: 12px; }}
  .totals .row.main {{ border-top: 2px solid #{GREEN}; margin-top: 6px; padding-top: 10px;
    font-size: 18px; font-weight: 800; color: #{GREEN_DARK}; }}
  .totals .row.save {{ color: #{GREEN}; font-weight: 700; }}
  footer {{ margin-top: 26px; color: #8a98a8; font-size: 11px; border-top: 1px solid #e3e9ef; padding-top: 10px; }}
</style></head>
<body>
  <header>
    <h1>Devis Cloud Temple — {e(m['project'])}</h1>
    <div class="meta">{date_line}<span>Projection : {e(period)}</span><span>Remise commerciale : {e(_pct(m['discount_percent']))}</span></div>
  </header>
  <table>
    <thead><tr>{headers}</tr></thead>
    <tbody>{''.join(rows)}</tbody>
  </table>
  <div class="totals">
    <div class="row muted"><span>Mensuel public</span><span class="v">{e(_eur(quote.monthly_public_total))}</span></div>
    <div class="row detail"><span>↳ dont remise catalogue</span><span>−{e(_eur(std_saving))}</span></div>
    <div class="row detail"><span>↳ dont remise commerciale</span><span>−{e(_eur(com_saving))}</span></div>
    <div class="row main"><span>Total mensuel remisé</span><span>{e(_eur(quote.monthly_discounted_total))}</span></div>
    <div class="row"><span>Projection {e(period)}</span><span>{e(_eur(quote.period_discounted_total))}</span></div>
    <div class="row"><span>Total à l'engagement</span><span>{e(_eur(quote.total_on_engagement))}</span></div>
    <div class="row save"><span>Économie sur {e(period)}</span><span>{e(_eur(quote.savings_total))}</span></div>
  </div>
  <footer>Tarifs HT en euros · remise standard catalogue appliquée automatiquement · Cloud Temple</footer>
</body></html>"""


# --------------------------------------------------------------------------- #
# PDF (reportlab)
# --------------------------------------------------------------------------- #
def quote_to_pdf(quote: QuoteResponse, meta: Optional[dict] = None) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    m = _meta(quote, meta)
    green = colors.HexColor("#16A06A")
    green_dark = colors.HexColor("#0F7A50")
    sub_bg = colors.HexColor("#E8F4EE")
    period = _period_label(m["period_months"])

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("t", parent=styles["Title"], textColor=green_dark, fontSize=18, alignment=0)
    meta_style = ParagraphStyle("m", parent=styles["Normal"], textColor=colors.HexColor("#5B6B7D"), fontSize=9)
    cell = ParagraphStyle("c", parent=styles["Normal"], fontSize=8, leading=10)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=14 * mm, rightMargin=14 * mm, topMargin=14 * mm, bottomMargin=14 * mm,
        title=f"Devis Cloud Temple — {m['project']}",
    )

    story: list[Any] = [
        Paragraph(f"Devis Cloud Temple — {_html.escape(m['project'])}", title_style),
        Spacer(1, 3),
    ]
    info = f"Projection : {period} &nbsp;&nbsp; Remise commerciale : {_pct(m['discount_percent'])}"
    if m["date"]:
        info = f"{m['date']} &nbsp;&nbsp; " + info
    story += [Paragraph(info, meta_style), Spacer(1, 8)]

    data = [[label for _, label in COLUMNS]]
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), green),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LINEBELOW", (0, 0), (-1, -1), 0.4, colors.HexColor("#E3E9EF")),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]

    r = 1
    for cat, lines in _grouped(quote).items():
        data.append([cat, "", "", "", "", "", "", "", ""])
        style_cmds += [
            ("SPAN", (0, r), (-1, r)),
            ("BACKGROUND", (0, r), (-1, r), sub_bg),
            ("TEXTCOLOR", (0, r), (0, r), green_dark),
            ("FONTNAME", (0, r), (0, r), "Helvetica-Bold"),
        ]
        r += 1
        subtotal = 0.0
        for l in lines:
            eng = f"{l.engagement_months} mois" if l.engagement_months > 1 else "—"
            data.append([
                Paragraph(_html.escape(l.name), cell),
                Paragraph(f"<font size=6>{_html.escape(l.sku)}</font>", cell),
                _qty(l.quantity),
                _eur(l.public_unit_price),
                _pct(l.standard_discount_percent),
                _eur(l.discounted_unit_price),
                _eur(l.monthly_total),
                eng,
                _eur(l.engagement_total),
            ])
            subtotal += l.monthly_total
            r += 1
        data.append(["", "", "", "", "", f"Sous-total {cat}", _eur(subtotal), "", ""])
        style_cmds += [
            ("FONTNAME", (5, r), (6, r), "Helvetica-Bold"),
            ("LINEABOVE", (5, r), (6, r), 0.4, green),
        ]
        r += 1

    col_widths = [62 * mm, 42 * mm, 12 * mm, 22 * mm, 18 * mm, 22 * mm, 24 * mm, 18 * mm, 28 * mm]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle(style_cmds))
    story.append(table)

    # Bloc totaux
    std_saving, com_saving = _savings_breakdown(quote)
    story.append(Spacer(1, 10))
    tot_rows = [
        ["Mensuel public", _eur(quote.monthly_public_total)],
        ["dont remise catalogue", "−" + _eur(std_saving)],
        ["dont remise commerciale", "−" + _eur(com_saving)],
        ["Total mensuel remisé", _eur(quote.monthly_discounted_total)],
        [f"Projection {period}", _eur(quote.period_discounted_total)],
        ["Total à l'engagement", _eur(quote.total_on_engagement)],
        [f"Économie sur {period}", _eur(quote.savings_total)],
    ]
    tot = Table(tot_rows, colWidths=[55 * mm, 35 * mm], hAlign="RIGHT")
    tot.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("TEXTCOLOR", (0, 1), (-1, 2), colors.HexColor("#5B6B7D")),
        ("LINEABOVE", (0, 3), (-1, 3), 1, green),
        ("FONTNAME", (0, 3), (-1, 3), "Helvetica-Bold"),
        ("FONTSIZE", (0, 3), (-1, 3), 12),
        ("TEXTCOLOR", (0, 3), (-1, 3), green_dark),
        ("TEXTCOLOR", (0, 6), (-1, 6), green),
        ("FONTNAME", (0, 6), (-1, 6), "Helvetica-Bold"),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    story.append(tot)

    doc.build(story)
    return buf.getvalue()


def render_quote(quote: QuoteResponse, fmt: str, meta: Optional[dict] = None) -> tuple[bytes, str, str]:
    """Retourne (contenu, content_type, extension) pour le format demandé."""
    fmt = (fmt or "").lower()
    if fmt == "xlsx":
        return (
            quote_to_xlsx(quote, meta),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "xlsx",
        )
    if fmt == "pdf":
        return quote_to_pdf(quote, meta), "application/pdf", "pdf"
    if fmt == "html":
        return quote_to_html(quote, meta).encode("utf-8"), "text/html; charset=utf-8", "html"
    raise ValueError(f"Format d'export non supporté : {fmt}")
